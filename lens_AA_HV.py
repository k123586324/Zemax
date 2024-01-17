import numpy as np
import openpyxl

def cal_txt(txt_path):
    text = []
    with open(txt_path, 'r', encoding='UTF-16') as f:
        for line in f:
            text.append(line)

    #CT
    for i in text:
        if "Field: 0.0000, 0.0000 mm" in i:
            CT_start=text.index(i)+2
            CT_end=CT_start+200
    CT_array=[[0]*3 for i in range(201)]
    j=0
    for i in range(CT_start,CT_end+1):
        CT_array[j][0] = text[i].split()[0]
        CT_array[j][1] = text[i].split()[1]
        CT_array[j][2] = text[i].split()[2]
        j+=1

    #UR
    for i in text:
        if "Field: 2.3040, 1.7280 mm" in i:
            UR_start=text.index(i)+2
            UR_end=UR_start+200
    UR_array=[[0]*3 for i in range(201)]
    j=0
    for i in range(UR_start,UR_end+1):
        UR_array[j][0] = text[i].split()[0]
        UR_array[j][1] = text[i].split()[1]
        UR_array[j][2] = text[i].split()[2]
        j+=1

    #UL
    for i in text:
        if "Field: -2.3040, 1.7280 mm" in i:
            UL_start=text.index(i)+2
            UL_end=UL_start+200
    UL_array=[[0]*3 for i in range(201)]
    j=0
    for i in range(UL_start,UL_end+1):
        UL_array[j][0] = text[i].split()[0]
        UL_array[j][1] = text[i].split()[1]
        UL_array[j][2] = text[i].split()[2]
        j+=1

    #BL
    for i in text:
        if "Field: -2.3040, -1.7280 mm" in i:
            BL_start=text.index(i)+2
            BL_end=BL_start+200
    BL_array=[[0]*3 for i in range(201)]
    j=0
    for i in range(BL_start,BL_end+1):
        BL_array[j][0] = text[i].split()[0]
        BL_array[j][1] = text[i].split()[1]
        BL_array[j][2] = text[i].split()[2]
        j+=1

    #BR
    for i in text:
        if "Field: 2.3040, -1.7280 mm" in i:
            BR_start=text.index(i)+2
            BR_end=BR_start+200
    BR_array=[[0]*3 for i in range(201)]
    j=0
    for i in range(BR_start,BR_end+1):
        BR_array[j][0] = text[i].split()[0]
        BR_array[j][1] = text[i].split()[1]
        BR_array[j][2] = text[i].split()[2]
        j+=1


    #計算中心
    CT_tan_max=0
    CT_tan_max_index=0
    for i in range(0,201):
        if float(CT_array[i][1])>=CT_tan_max:
            CT_tan_max=float(CT_array[i][1])
            CT_tan_max_index = i
    CT_sag_max=0
    CT_sag_max_index=0
    for i in range(0,201):
        if float(CT_array[i][2])>=CT_sag_max:
            CT_sag_max=float(CT_array[i][2])
            CT_sag_max_index = i
    CT_tan_max_MTF=float(CT_array[CT_tan_max_index][1])
    CT_sag_max_MTF=float(CT_array[CT_sag_max_index][2])
    CT_st_distance=abs(float(CT_array[CT_tan_max_index][0])-float(CT_array[CT_sag_max_index][0]))*1000
    #print(CT_tan_max_MTF,",",CT_sag_max_MTF,",",CT_st_distance)

    #計算UR
    UR_tan_max=0
    UR_tan_max_index=0
    for i in range(0,201):
        if float(UR_array[i][1])>=UR_tan_max:
            UR_tan_max=float(UR_array[i][1])
            UR_tan_max_index = i
    UR_sag_max=0
    UR_sag_max_index=0
    for i in range(0,201):
        if float(UR_array[i][2])>=UR_sag_max:
            UR_sag_max=float(UR_array[i][2])
            UR_sag_max_index = i
    UR_tan_max_MTF=float(UR_array[UR_tan_max_index][1])
    UR_sag_max_MTF=float(UR_array[UR_sag_max_index][2])
    UR_st_distance=abs(float(UR_array[UR_tan_max_index][0])-float(UR_array[UR_sag_max_index][0]))*1000
    #print(UR_tan_max_MTF,",",UR_sag_max_MTF,",",UR_st_distance)

    #計算UL
    UL_tan_max=0
    UL_tan_max_index=0
    for i in range(0,201):
        if float(UL_array[i][1])>=UL_tan_max:
            UL_tan_max=float(UL_array[i][1])
            UL_tan_max_index = i
    UL_sag_max=0
    UL_sag_max_index=0
    for i in range(0,201):
        if float(UL_array[i][2])>=UL_sag_max:
            UL_sag_max=float(UL_array[i][2])
            UL_sag_max_index = i
    UL_tan_max_MTF=float(UL_array[UL_tan_max_index][1])
    UL_sag_max_MTF=float(UL_array[UL_sag_max_index][2])
    UL_st_distance=abs(float(UL_array[UL_tan_max_index][0])-float(UL_array[UL_sag_max_index][0]))*1000
    #print(UL_tan_max_MTF,",",UL_sag_max_MTF,",",UL_st_distance)

    #計算BL
    BL_tan_max=0
    BL_tan_max_index=0
    for i in range(0,201):
        if float(BL_array[i][1])>=BL_tan_max:
            BL_tan_max=float(BL_array[i][1])
            BL_tan_max_index = i
    BL_sag_max=0
    BL_sag_max_index=0
    for i in range(0,201):
        if float(BL_array[i][2])>=BL_sag_max:
            BL_sag_max=float(BL_array[i][2])
            BL_sag_max_index = i
    BL_tan_max_MTF=float(BL_array[BL_tan_max_index][1])
    BL_sag_max_MTF=float(BL_array[BL_sag_max_index][2])
    BL_st_distance=abs(float(BL_array[BL_tan_max_index][0])-float(BL_array[BL_sag_max_index][0]))*1000
    #print(BL_tan_max_MTF,",",BL_sag_max_MTF,",",BL_st_distance)

    #計算BR
    BR_tan_max=0
    BR_tan_max_index=0
    for i in range(0,201):
        if float(BR_array[i][1])>=BR_tan_max:
            BR_tan_max=float(BR_array[i][1])
            BR_tan_max_index = i
    BR_sag_max=0
    BR_sag_max_index=0
    for i in range(0,201):
        if float(BR_array[i][2])>=BR_sag_max:
            BR_sag_max=float(BR_array[i][2])
            BR_sag_max_index = i
    BR_tan_max_MTF=float(BR_array[BR_tan_max_index][1])
    BR_sag_max_MTF=float(BR_array[BR_sag_max_index][2])
    BR_st_distance=abs(float(BR_array[BR_tan_max_index][0])-float(BR_array[BR_sag_max_index][0]))*1000
    #print(BR_tan_max_MTF,",",BR_sag_max_MTF,",",BR_st_distance)

    UR_tan_avg=UR_tan_max_MTF-(UL_tan_max_MTF+BL_tan_max_MTF+BR_tan_max_MTF)/3
    UR_sag_avg=UR_sag_max_MTF-(UL_sag_max_MTF+BL_sag_max_MTF+BR_sag_max_MTF)/3
    UL_tan_avg=UL_tan_max_MTF-(UR_tan_max_MTF+BL_tan_max_MTF+BR_tan_max_MTF)/3
    UL_sag_avg=UL_sag_max_MTF-(UR_sag_max_MTF+BL_sag_max_MTF+BR_sag_max_MTF)/3
    BL_tan_avg=BL_tan_max_MTF-(UL_tan_max_MTF+UR_tan_max_MTF+BR_tan_max_MTF)/3
    BL_sag_avg=BL_sag_max_MTF-(UL_sag_max_MTF+UR_sag_max_MTF+BR_sag_max_MTF)/3
    BR_tan_avg=BR_tan_max_MTF-(UL_tan_max_MTF+BL_tan_max_MTF+UR_tan_max_MTF)/3
    BR_sag_avg=BR_sag_max_MTF-(UL_sag_max_MTF+BL_sag_max_MTF+UR_sag_max_MTF)/3

    output=[0]*16
    output[0]=CT_tan_max_MTF
    output[1]=CT_sag_max_MTF
    output[2]=CT_st_distance
    output[3] = ""
    output[4]=UR_tan_avg
    output[5]=UR_sag_avg
    output[6]=UL_tan_avg
    output[7]=UL_sag_avg
    output[8]=BL_tan_avg
    output[9]=BL_sag_avg
    output[10]=BR_tan_avg
    output[11]=BR_sag_avg
    output[12]=UR_st_distance
    output[13]=UL_st_distance
    output[14]=BL_st_distance
    output[15]=BR_st_distance
    return output




workbook = openpyxl.load_workbook('AA_算法.xlsx')
sheet = workbook.worksheets[0]

excel_initial_index=156
excel_end_index=186
txt_initial_index=-75
txt_step=5

for i in range(excel_initial_index, excel_end_index+1):
    txt_index=txt_initial_index+(i-excel_initial_index)*txt_step
    data = cal_txt(f'C:\\Users\\Alfieliu\\Desktop\\data\\G2DY\\{txt_index}.0000.txt')
    for j in range(4, sheet.max_column+1):
        sheet.cell(row = i, column=j).value=data[j-4]
workbook.save('AA_算法.xlsx')





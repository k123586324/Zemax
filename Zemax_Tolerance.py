import numpy as np
from openpyxl.chart import ScatterChart, Reference, Series
import openpyxl
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk


def zmx_tol(path):
    try:
        # 開啟讀取檔案
        max_row = 0
        fields = 0
        with open(path, 'r', encoding='UTF-16 LE') as fh:
            tmp = fh.read()
            onelinelist = tmp.split('\n')

        # 公差資料儲存至2D矩陣
        for i in range(len(onelinelist)):
            # print(onelinelist[i])
            if "Field by Field Nominal Values:" in onelinelist[i]:
                while onelinelist[i + fields + 3] != "":
                    fields += 1
            if "Number of trials" in onelinelist[i]:
                max_row = int(onelinelist[i].split(":")[1])
                tol_data = np.zeros((max_row, fields))
                for y in range(max_row):
                    tmp = onelinelist[i + 6 + y].split()
                    for x in range(fields):
                        tol_data[y][x] = tmp[x + 3]
                break

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = '工作表1'
        for x in range(fields):
            sheet.cell(row=1, column=2 + x, value=f'Field {x + 1}')
        sheet.cell(row=1, column=fields + 2, value="Percentage")
        for y in range(max_row):
            sheet.cell(row=2 + y, column=1, value=y + 1)
            sheet.cell(row=2 + y, column=fields + 2, value=((max_row - y) / max_row) * 100)
        for x in range(fields):
            tmp_array = np.zeros(max_row)
            for y in range(max_row):
                tmp_array[y] = tol_data[y][x]
            # tmp_array=abs(np.sort(-tmp_array))
            tmp_array = np.sort(tmp_array)
            for y in range(max_row):
                sheet.cell(row=2 + y, column=2 + x, value=tmp_array[y])

        # 創建圖表
        chart = ScatterChart()
        chart.title = "Zemax Tolerance"
        chart.style = 42
        chart.x_axis.title = 'MTF'
        chart.y_axis.title = 'Percentage'
        chart.width = 20
        chart.height = 10
        # 選擇數據範圍
        for i in range(2, fields + 2):
            x_values = Reference(sheet, min_col=i, min_row=2, max_row=max_row + 1)
            y_values = Reference(sheet, min_col=fields + 2, min_row=2, max_row=max_row + 1)
            series = Series(y_values, x_values, title=sheet.cell(row=1, column=i).value)
            series.smooth = True
            chart.series.append(series)
        chart.x_axis.scaling.max = 1
        chart.x_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100
        chart.y_axis.scaling.min = 0
        chart.x_axis.scaling.orientation = "maxMin"
        sheet.add_chart(chart, "A1")

        workbook.save(r'Zemax_Tolerance.xlsx')
        messagebox.showinfo(title="訊息", message="完成")
    except:
        messagebox.showinfo(title="訊息", message="失敗")


app = tk.Tk()
app.title("Zemax_Tolerance")
app.geometry('300x100+500+250')
app.resizable(False, False)

label_files = tk.Label(app, text=f'Version 1.0 by Alfie')
label_files.place(x=180, y=75)


def main():
    path = filedialog.askopenfilename()
    zmx_tol(path)


button1 = tk.Button(app, text="選擇檔案", command=main)
button1.place(width=100,height=50,x=100, y=20)


def on_closing():
    app.destroy()


app.protocol('WM_DELETE_WINDOW', on_closing)
app.mainloop()
